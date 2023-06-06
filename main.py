import openpyxl
import json


# Открываем xlsx-файл
def get_input_row(filename):
    wb = openpyxl.load_workbook(filename)

    # Получаем активный лист
    sheet = wb.active

    # Читаем и выводим каждую строку
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)


def get_main_image(images):
    space_index = images.find(" ")
    if space_index != -1:
        return images[:space_index]
    else:
        if len(images) > 5:
            return images
        else:
            return ""


def get_additional_images(images):
    space_index = images.find(" ")
    if space_index != -1:
        return images[space_index + 1:]
    else:
        return ""


def get_weight(input_row):
    weight = input_row[1]
    if not weight:
        return '150'
    else:
        return str(weight).strip()


def get_basic_row(row, input_row, commercial_type):
    row.append("")
    row.append(input_row[0])  # артикул
    row.append(get_title(input_row))  # название товара
    row.append(input_row[5])
    row.append(input_row[6])
    row.append("Не облагается")
    row.append("")
    row.append(commercial_type)
    row.append("")
    row.append(input_row[1])
    row.append(input_row[3])
    row.append(input_row[4])
    row.append(input_row[2])
    row.append(get_main_image(input_row[10]))
    row.append(get_additional_images(input_row[10]))


def get_count_in_pack(input_row):
    input_count = input_row[12]
    if input_count == "" or not input_count:
        return "1"
    else:
        return input_count


def get_color_name(input_row):
    return input_row[13]


def get_brand(input_row):
    return input_row[11]


def title_normalize(title):
    if not title or title == "":
        return ''
    else:
        is_contains_big_numbers = False
        title_arr = title.split(',')
        new_title_arr = []
        for item in title_arr:
            if item.strip().isdigit():
                if int(item) > 5000:
                    is_contains_big_numbers = True
            else:
                new_title_arr.append(item)
        if is_contains_big_numbers:
            return ' '.join(new_title_arr).strip()
        else:
            return str(title).replace(',', ' ').strip()


def get_title(input_row):
    title = input_row[8]
    return title_normalize(title)


def get_type(commercial_type):
    categories_match_file = open("types.json", 'r', encoding='utf-8')
    categories_match_dict = json.load(categories_match_file)
    return categories_match_dict.get(commercial_type, "")


def get_description(input_row):
    return input_row[9]


def get_stationery_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Канцелярия """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append(get_count_in_pack(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    row.extend(['', ''])
    row.append("")
    return row


def get_sheets_count(input_row):
    sheets_count = input_row[14]
    if sheets_count == "" or not sheets_count:
        return '100'
    else:
        return sheets_count


def get_density(input_row):
    density = input_row[15]
    if density == "" or not density:
        return '60'
    else:
        return density


def get_paper_products_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Канцелярские товары > Бумажная продукция """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_sheets_count(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    row.append("")
    row.append("")
    row.append("")
    row.append(get_density(input_row))  # плотность бумаги
    return row


def get_folders_files_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Канцелярские товары > Папки и файлы """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.append("")
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_medical_devices_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Аптека > Медицинские изделия """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append("")
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(['', '', '', '', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_size(input_row):
    size = input_row[16]
    if not size:
        return ""
    else:
        return size


def get_expiration_date():
    return '730'


def get_medical_supplies_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Аптека > Медицинские расходные материалы """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append("")
    row.append("")  # в упаковке
    row.append("")  # в товаре
    row.append("")
    row.append(get_type(commercial_type))
    row.append(get_expiration_date())  # срок годности
    row.extend(['', ''])
    row.append(get_description(input_row))
    return row


def get_cleaning_products_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Бытовая химия > Моющие и чистящие средства """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_expiration_date())  # срок годности
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(["", ""])
    row.append("")
    row.extend(['', '', '', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_air_freshener_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Бытовая химия > Освежители воздуха """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append("")
    row.append("")  # в упаковке
    row.append(get_expiration_date())  # срок годности
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.append("")
    row.extend(['', '', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_bags_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Галантерея и украшения > Сумка """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append(get_title(input_row))
    row.append("")
    row.extend(['', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_food_accessories_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Дом и сад > Аксессуары для приготовления пищи """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_inventory_for_home_row(input_row, commercial_type):
    """ Генерирует строку для записи в Дом и сад > Инвентарь для дома """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_inventory_for_cleaning_row(input_row, commercial_type):
    """ Генерирует строку для записи в Дом и сад > Инвентарь для уборки """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_disposable_tableware_row(input_row, commercial_type):
    """ Генерирует строку для записи в Дом и сад > Одноразовая посуда """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_dishes_row(input_row, commercial_type):
    """ Генерирует строку для записи в Дом и сад > Столовая посуда """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append("")
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_things_storage_row(input_row, commercial_type):
    """ Генерирует строку для записи в Дом и сад > Хранение вещей """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.extend(["", "", "", ""])
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_paper_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Бумага """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_sheets_count(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_demonstration_boards_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Демонстрационные доски """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_child_bags_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Детские рюкзаки, ранцы, сумки """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(['', '', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_glue_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Краска, клей """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append("")
    row.append("")
    row.append(get_expiration_date())  # срок годности
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_pencil_box_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Пенал """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append("")
    row.append("")
    row.append(get_type(commercial_type))
    row.append("Каркасный пенал")  # вид пенала ,
    row.append("1")  # кол-во отделений
    row.append("Нет")  # наполнение пенала
    row.append("")
    row.append("")
    row.extend(['', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_seal_and_stamp_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Печати и штампы """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.append('')
    row.extend(['', '', ''])
    row.append(get_description(input_row))
    return row


def get_writing_materials_row(input_row, commercial_type):
    """ Генерирует строку для записи в Канцелярские товары > Письменные принадлежности """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_color_name(input_row))
    row.append("")
    row.append(get_type(commercial_type))
    row.append("")
    row.append("")
    row.extend(['', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_personal_hygiene_row(input_row, commercial_type):
    """ Генерирует строку для записи в Красота и Гигиена > Товары личной гигиены """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_title(input_row))
    row.append(get_brand(input_row))
    row.append("")
    row.append("")
    row.append("")
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append(get_expiration_date())  # срок годности
    row.extend(['', ''])
    row.append(get_description(input_row))
    row.extend(['', '', ''])
    row.append("")
    return row


def get_clothes_row(input_row, commercial_type):
    """ Генерирует строку для записи в Одежда > Одежда """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append("универсальный")
    row.append(get_size(input_row))
    row.append(get_color_name(input_row))
    row.append(get_type(commercial_type))
    row.append("Мужской")
    row.extend(['', '', '', '', '', ''])
    row.append("")
    row.append('')
    row.append(get_description(input_row))
    return row


def get_juse_drinks_row(input_row, commercial_type):
    """ Генерирует строку для записи в Продукты питания > Соки, воды, напитки """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append(get_count_in_pack(input_row))
    row.append(get_weight(input_row))
    row.append(get_type(commercial_type))
    row.append("")  # максимальная температура
    row.append("")  # минимальная температура
    row.append(get_expiration_date())  # срок годности
    row.append("")  # условия хранения как на упаковке
    row.append("")  # состав
    row.extend(['', ''])
    row.append(get_description(input_row))
    row.extend(['', '', '', '', '', '', '', '', '', ''])
    row.append("")
    return row


def get_bread_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Продукты питания > Хлеб и кондитерские изделия """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append(get_weight(input_row))
    row.append(get_count_in_pack(input_row))
    row.append(get_type(commercial_type))
    row.append("")  # максимальная температура
    row.append("")  # минимальная температура
    row.append(get_expiration_date())  # срок годности
    row.append("")  # условия хранения как на упаковке
    row.append("")  # состав
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    row.extend(['', '', '', '', '', '', '', '', '', '', '', ''])
    row.append("")
    return row


def get_cokol_lamp(input_row):
    cokol = input_row[17]
    if not cokol:
        return 'E27'
    else:
        return str(cokol).strip()


def get_lamp_row(input_row, commercial_type):
    """ Генерирует строку для записи в Строительство и ремонт > Лампочка """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_count_in_pack(input_row))
    row.append(get_cokol_lamp(input_row))  # тип цоколя
    row.append("")  # мощность, ВТ
    row.append(get_type(commercial_type))
    row.extend(['', '', ''])
    row.append("")
    row.extend(['', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_big_lamp_row(input_row, commercial_type):
    """ Генерирует строку для записи в Строительство и ремонт > Светильник """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.extend(["", "", ""])
    row.append(get_type(commercial_type))
    row.extend(['', '', '', '', '', '', '', ''])
    row.append(get_description(input_row))
    row.extend(['', ''])
    row.append("")
    return row


def get_fire_fighting_row(input_row, commercial_type):
    """ Генерирует строку для записи в Строительство и ремонт > Средства защиты и пожаротушения """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.extend(["", ""])
    row.append(get_type(commercial_type))
    row.extend(['', ''])
    row.append("")
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_children_creativity_row(input_row, commercial_type):
    """ Генерирует строку для записи в Хобби и творчество > Детское творчество и развитие """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.extend(["", ""])
    row.append(get_type(commercial_type))
    row.extend([''])
    row.append("")
    row.extend(['', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_set_for_creativity_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Хобби и творчество > Набор для рукоделия, творчества """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.extend(["", ""])
    row.append(get_type(commercial_type))
    row.extend([''])
    row.append("")
    row.extend(['', '', '', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_office_equipment_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Электроника > Офисная техника """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.extend([""])
    row.append(get_type(commercial_type))
    row.extend(['', '', '', ''])
    row.append(get_description(input_row))
    return row


def get_computer_bag_row(input_row, commercial_type):
    """ Генерирует строку для записи в  Электроника > Рюкзаки, чехлы, сумки """
    row = []
    get_basic_row(row, input_row, commercial_type)
    row.extend(["", ""])
    row.append(get_brand(input_row))
    row.append(get_title(input_row))
    row.append("")
    row.append(get_type(commercial_type))
    row.extend([''])
    row.append("")
    row.extend(['', ''])
    row.append(get_description(input_row))
    return row


def get_first_dict_value(dict):
    if type(dict) == str:
        return 1
    for i in dict:
        if get_first_dict_value(dict[i]) == 1:
            return dict[i]


def get_ozon_category(initial_category: str):
    categories_match_file = open("categories.json", 'r', encoding='utf-8')
    categories_match_dict = json.load(categories_match_file)
    categories_match_file.close()
    initial_category = initial_category.split("/")
    our_category = categories_match_dict
    for category in initial_category:
        if category in our_category:
            our_category = our_category[category]
        else:
            return False
    if type(our_category) == str:
        return our_category
    elif type(our_category) == dict:
        return get_first_dict_value(our_category)


def write_row(filename, row):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.append(row)
    wb.save(filename)


def main():
    wb = openpyxl.load_workbook("input.xlsx")
    counter = 0
    # Получаем активный лист
    sheet = wb.active
    # Читаем и выводим каждую строку
    chancellery_categories = [i.strip() for i in
                              open("categories_files/chancellery.txt", 'r', encoding='utf-8').readlines()]
    paper_products_categories = [i.strip() for i in
                                 open("categories_files/paper_products_categories.txt", 'r',
                                      encoding='utf-8').readlines()]
    medical_devices_categories = [i.strip() for i in
                                  open("categories_files/medical_devices_categories.txt", 'r',
                                       encoding='utf-8').readlines()]
    medical_supplies_categories = [i.strip() for i in
                                   open("categories_files/medical_supplies_categories.txt", 'r',
                                        encoding='utf-8').readlines()]
    cleaning_products_categories = [i.strip() for i in
                                    open("categories_files/cleaning_products_categories.txt", 'r',
                                         encoding='utf-8').readlines()]
    air_freshener_categories = [i.strip() for i in
                                open("categories_files/air_freshener_categories.txt", 'r',
                                     encoding='utf-8').readlines()]
    inventory_for_home_categories = [i.strip() for i in
                                     open("categories_files/inventory_for_home_categories.txt", 'r',
                                          encoding='utf-8').readlines()]
    inventory_for_cleaning_categories = [i.strip() for i in
                                         open("categories_files/inventory_for_cleaning_categories.txt", 'r',
                                              encoding='utf-8').readlines()]
    dishes_categories = [i.strip() for i in
                         open("categories_files/dishes_categories.txt", 'r',
                              encoding='utf-8').readlines()]
    things_storage_categories = [i.strip() for i in
                                 open("categories_files/things_storage_categories.txt", 'r',
                                      encoding='utf-8').readlines()]

    paper_categories = [i.strip() for i in
                        open("categories_files/paper_categories.txt", 'r',
                             encoding='utf-8').readlines()]
    demonstration_boards_categories = [i.strip() for i in
                                       open("categories_files/demonstration_boards_categories.txt", 'r',
                                            encoding='utf-8').readlines()]
    child_bags_categories = [i.strip() for i in
                             open("categories_files/child_bags_categories.txt", 'r',
                                  encoding='utf-8').readlines()]
    folders_files_categories = [i.strip() for i in
                                open("categories_files/folders_files_categories.txt", 'r',
                                     encoding='utf-8').readlines()]
    seal_and_stamp_categories = [i.strip() for i in
                                 open("categories_files/seal_and_stamp_categories.txt", 'r',
                                      encoding='utf-8').readlines()]
    writing_materials_categories = [i.strip() for i in
                                    open("categories_files/writing_materials_categories.txt", 'r',
                                         encoding='utf-8').readlines()]
    personal_hygiene_categories = [i.strip() for i in
                                   open("categories_files/personal_hygiene_categories.txt", 'r',
                                        encoding='utf-8').readlines()]
    juse_drinks_categories = [i.strip() for i in
                              open("categories_files/juse_drinks_categories.txt", 'r',
                                   encoding='utf-8').readlines()]
    bread_categories = [i.strip() for i in
                        open("categories_files/bread_categories.txt", 'r',
                             encoding='utf-8').readlines()]
    fire_fighting_categories = [i.strip() for i in
                                open("categories_files/fire_fighting_categories.txt", 'r',
                                     encoding='utf-8').readlines()]
    children_creativity_categories = [i.strip() for i in
                                      open("categories_files/children_creativity_categories.txt", 'r',
                                           encoding='utf-8').readlines()]
    set_for_creativity_categories = [i.strip() for i in
                                     open("categories_files/set_for_creativity_categories.txt", 'r',
                                          encoding='utf-8').readlines()]
    office_equipment_categories = [i.strip() for i in
                                   open("categories_files/office_equipment_categories.txt", 'r',
                                        encoding='utf-8').readlines()]
    for input_row in sheet.iter_rows(values_only=True):
        try:
            ozon_category = get_ozon_category(input_row[7])
            if ozon_category in chancellery_categories:
                counter += 1
                row = get_stationery_row(input_row, ozon_category)
                write_row("output/chancellery_output.xlsx", row)
            elif ozon_category in paper_products_categories:
                counter += 1
                row = get_paper_products_row(input_row, ozon_category)
                write_row("output/paper_products_output.xlsx", row)
            elif ozon_category in medical_devices_categories:
                counter += 1
                row = get_medical_devices_row(input_row, ozon_category)
                write_row("output/medical_devices_output.xlsx", row)
            elif ozon_category in medical_supplies_categories:
                counter += 1
                row = get_medical_supplies_row(input_row, ozon_category)
                write_row("output/medical_supplies_output.xlsx", row)
            elif ozon_category in cleaning_products_categories:
                counter += 1
                row = get_cleaning_products_row(input_row, ozon_category)
                write_row("output/cleaning_products_output.xlsx", row)
            elif ozon_category in air_freshener_categories:
                counter += 1
                row = get_air_freshener_row(input_row, ozon_category)
                write_row("output/air_freshener_output.xlsx", row)
            elif ozon_category == "Женская Сумка-шоппер":
                counter += 1
                row = get_bags_row(input_row, ozon_category)
                write_row("output/bags_output.xlsx", row)
            elif ozon_category == "Фольга, пленка, бумага для выпечки, пакеты для запекания":
                counter += 1
                row = get_food_accessories_row(input_row, ozon_category)
                write_row("output/food_accessories_output.xlsx", row)
            elif ozon_category in inventory_for_home_categories:
                counter += 1
                row = get_inventory_for_home_row(input_row, ozon_category)
                write_row("output/inventory_for_home_output.xlsx", row)
            elif ozon_category in inventory_for_cleaning_categories:
                counter += 1
                row = get_inventory_for_cleaning_row(input_row, ozon_category)
                write_row("output/inventory_for_cleaning_output.xlsx", row)
            elif ozon_category == "Посуда одноразовая":
                counter += 1
                row = get_disposable_tableware_row(input_row, ozon_category)
                write_row("output/disposable_tableware_output.xlsx", row)
            elif ozon_category in dishes_categories:
                counter += 1
                row = get_dishes_row(input_row, ozon_category)
                write_row("output/dishes_output.xlsx", row)
            elif ozon_category in things_storage_categories:
                counter += 1
                row = get_things_storage_row(input_row, ozon_category)
                write_row("output/things_storage_output.xlsx", row)
            elif ozon_category in paper_categories:
                counter += 1
                row = get_paper_row(input_row, ozon_category)
                write_row("output/paper_output.xlsx", row)
            elif ozon_category in demonstration_boards_categories:
                counter += 1
                row = get_demonstration_boards_row(input_row, ozon_category)
                write_row("output/demonstration_boards_output.xlsx", row)
            elif ozon_category in child_bags_categories:
                counter += 1
                row = get_child_bags_row(input_row, ozon_category)
                write_row("output/child_bags_output.xlsx", row)
            elif ozon_category == "Клей канцелярский":
                counter += 1
                row = get_glue_row(input_row, ozon_category)
                write_row("output/glue_output.xlsx", row)
            elif ozon_category in folders_files_categories:
                counter += 1
                row = get_folders_files_row(input_row, ozon_category)
                write_row("output/folders_files_output.xlsx", row)
            elif ozon_category == "Пенал без наполнения":
                counter += 1
                row = get_pencil_box_row(input_row, ozon_category)
                write_row("output/pencil_box_output.xlsx", row)
            elif ozon_category in seal_and_stamp_categories:
                counter += 1
                row = get_seal_and_stamp_row(input_row, ozon_category)
                write_row("output/seal_and_stamp_output.xlsx", row)
            elif ozon_category in writing_materials_categories:
                counter += 1
                row = get_writing_materials_row(input_row, ozon_category)
                write_row("output/writing_materials_output.xlsx", row)
            elif ozon_category in personal_hygiene_categories:
                counter += 1
                row = get_personal_hygiene_row(input_row, ozon_category)
                write_row("output/personal_hygiene_output.xlsx", row)
            elif ozon_category == "Одежда медицинская одноразовая":
                counter += 1
                row = get_clothes_row(input_row, ozon_category)
                write_row("output/clothes_output.xlsx", row)
            elif ozon_category in juse_drinks_categories:
                counter += 1
                row = get_juse_drinks_row(input_row, ozon_category)
                write_row("output/juse_drinks_output.xlsx", row)
            elif ozon_category in bread_categories:
                counter += 1
                row = get_bread_row(input_row, ozon_category)
                write_row("output/bread_output.xlsx", row)
            elif ozon_category == "Лампочка":
                counter += 1
                row = get_lamp_row(input_row, ozon_category)
                write_row("output/lamp_output.xlsx", row)
            elif ozon_category == "Светильник настольный":
                counter += 1
                row = get_big_lamp_row(input_row, ozon_category)
                write_row("output/big_lamp_output.xlsx", row)
            elif ozon_category in fire_fighting_categories:
                counter += 1
                row = get_fire_fighting_row(input_row, ozon_category)
                write_row("output/fire_fighting_output.xlsx", row)
            elif ozon_category in children_creativity_categories:
                counter += 1
                row = get_children_creativity_row(input_row, ozon_category)
                write_row("output/children_creativity_output.xlsx", row)
            elif ozon_category in set_for_creativity_categories:
                counter += 1
                row = get_set_for_creativity_row(input_row, ozon_category)
                write_row("output/set_for_creativity_output.xlsx", row)
            elif ozon_category == "Рюкзак для ноутбука":
                counter += 1
                row = get_computer_bag_row(input_row, ozon_category)
                write_row("output/computer_bag_output.xlsx", row)
            elif ozon_category in office_equipment_categories:
                counter += 1
                row = get_office_equipment_row(input_row, ozon_category)
                write_row("output/office_equipment_output.xlsx", row)
            else:
                f = open("not_handler.txt", 'a', encoding='utf-8')
                f.write(input_row[0] + '\n')
                f.close()
            print(counter)
        except Exception as e:
            f1 = open("error_ids.txt", 'a', encoding='utf-8')
            f1.write(input_row[0] + '\n')
            f1.close()
            f = open("errors.txt", 'a', encoding='utf-8')
            f.write('\n\nstart_bag\n' + 'id:' + input_row[0] + '\n')
            f.write(e.__str__() + '\nendbag\n\n')
            f.close()


def output_files_create():
    output_filenames = ['chancellery_output.xlsx', 'paper_products_output.xlsx', 'medical_devices_output.xlsx',
                        'medical_supplies_output.xlsx', 'cleaning_products_output.xlsx', 'air_freshener_output.xlsx',
                        'bags_output.xlsx', 'food_accessories_output.xlsx', 'inventory_for_home_output.xlsx',
                        'inventory_for_cleaning_output.xlsx',
                        'disposable_tableware_output.xlsx', 'dishes_output.xlsx', 'things_storage_output.xlsx',
                        'paper_output.xlsx', 'demonstration_boards_output.xlsx', 'child_bags_output.xlsx',
                        'glue_output.xlsx', 'folders_files_output.xlsx', 'pencil_box_output.xlsx',
                        'seal_and_stamp_output.xlsx', 'writing_materials_output.xlsx',
                        'personal_hygiene_output.xlsx',
                        'clothes_output.xlsx', 'juse_drinks_output.xlsx', 'bread_output.xlsx', 'lamp_output.xlsx',
                        'fire_fighting_output.xlsx', 'children_creativity_output.xlsx',
                        'set_for_creativity_output.xlsx', 'computer_bag_output.xlsx', 'office_equipment_output.xlsx',
                        'big_lamp_output.xlsx']
    print(len(output_filenames))
    dir = "output/"
    for file in output_filenames:
        wb = openpyxl.Workbook()  # создание нового файла
        wb.save(dir + file)  # сохранение файла


if __name__ == "__main__":
    main()
    #output_files_create()
